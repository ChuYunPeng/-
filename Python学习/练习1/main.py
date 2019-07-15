# 读取excel数据
import openpyxl
wb = openpyxl.load_workbook('445874-4.xlsx')  # 打开工作簿
sheet = wb['Supplementary Table 4']  # 根据名称获取第四个工作表中的数据
max_row = 209  # 输入最大行数
max_column = 13  # 输入最大列数
SNPdata={}  #定义一个字典
for i in range(0,max_column):
    name = sheet.cell(row=1,column=i+1). value  # 获取列名称
    SNPdata.setdefault(name,[])  # 字典中，列名称作为键值，其对应值初始化为空列表
    for row in range(2,max_row+1):
        SNPdata[name].append(sheet.cell(row=row,column=i+1). value)  # 将数据导入对应列表

# 打开NCBI并搜寻基因名称
import requests,bs4
for n in range(0,max_row-1):
    SNPname=str(SNPdata['SNP'][n])
    res = requests.get('https://www.ncbi.nlm.nih.gov/snp/'+SNPname)  # 下载网站
    NCBI = bs4.BeautifulSoup(res.content,features='lxml')
    elems = NCBI.select('#main_content > main > div.summary-box.usa-grid-full > dl:nth-child(2) > dd:nth-child(4) > span')  # 搜寻指定元素
    if len(elems) == 1:  # 若搜寻成功则输出名字，否则输出none
        genename = elems[0].getText().split()[0]
    else:
        genename = 'None'
    SNPdata.setdefault('gene',[])
    SNPdata['gene'].append(genename)
    print (n,SNPname,genename)

# 输出至excel中
new_wb = openpyxl.Workbook()  # 打开工作簿
new_ws = new_wb.active  # 打开当前工作表
new_ws.title = 'SNP&GENES'  # 更改工作表名称
new_name=['CHR','gene','SNP','P','OR','BP','SE','A1A2', 'FRQ_A', 'FRQ_U', 'INFO', 'Nca', 'Nco', 'Neff']  # 列名称
for i in range(0,14):  # 输入数据
    new_ws.cell(row=1,column=i+1).value = new_name[i]
    for row in range(2, max_row + 1):
        new_ws.cell(row=row,column=i+1).value = SNPdata[new_name[i]][row-2]
new_wb.save('SNP&GENES.xlsx')  # 保存工作簿